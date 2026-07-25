from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.security import hash_password
from app.models.user import User


def _clean(value) -> str:
    if value is None:
        return ""
    result = str(value).strip()
    return "" if result == "-" else result


def _employee_rows(seed_file: Path) -> dict[str, dict[str, str]]:
    workbook = load_workbook(seed_file, read_only=True, data_only=True)
    worksheet = workbook.active
    employees: dict[str, dict[str, str]] = {}

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        category, department, job_title, display_name, extension, email, *_ = row
        email = _clean(email).lower()
        if not email or "@" not in email:
            continue

        username = email.split("@", 1)[0]
        employees[username] = {
            "username": username,
            "display_name": _clean(display_name),
            "email": email,
            "category": _clean(category),
            "department": _clean(department),
            "job_title": _clean(job_title),
            "extension": _clean(extension),
        }

    workbook.close()
    return employees


def seed_users(db: Session) -> tuple[int, int]:
    """Create missing employees without ever resetting an existing password."""
    if not settings.USERS_SEED_ENABLED:
        return 0, 0

    seed_file = Path(settings.USERS_SEED_FILE)
    if not seed_file.is_absolute():
        backend_dir = Path(__file__).resolve().parents[2]
        seed_file = backend_dir / seed_file
    seed_file = seed_file.resolve()
    if not seed_file.exists():
        raise RuntimeError(f"Employee seed file was not found: {seed_file}")

    created = 0
    initialized = 0
    for employee in _employee_rows(seed_file).values():
        user = db.query(User).filter(User.username == employee["username"]).first()
        if user is None:
            user = User(
                **employee,
                password_hash=hash_password(settings.DEFAULT_USER_PASSWORD),
                must_change_password=True,
            )
            db.add(user)
            created += 1
            continue

        # Upgrade old locally stored users once, but never overwrite a password.
        if not user.password_hash:
            user.password_hash = hash_password(settings.DEFAULT_USER_PASSWORD)
            user.must_change_password = True
            initialized += 1

        # Organization-owned data follows the employee source. User-editable
        # display name and email are only filled when they are missing.
        user.category = employee["category"]
        user.department = employee["department"]
        user.job_title = employee["job_title"]
        user.extension = employee["extension"]
        if not user.display_name:
            user.display_name = employee["display_name"]
        if not user.email:
            user.email = employee["email"]

    db.commit()
    return created, initialized
